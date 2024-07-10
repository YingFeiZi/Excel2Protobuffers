import re
import config
import openpyxl
import sys
from pathlib import Path

cusrp = ['int32#&list', 'int32#|list', 'int32#|&list']
def ReplaceType(value):
    # if value == 'uint32list' or value == 'uint32' or value == 'sint32':
    if value == 'int32#&list':
        return 'pair#&list'
    if value == 'int32#|list':
        return 'pair#|list'
    if value == 'int32#|&list':
        return 'pair#|&list'
    return value

def HasHtml(html):
    # 安全地处理HTML，避免XSS攻击，确保替换的内容被正确转义
    def check_patterns(text):
        has_font_color =  bool(pattern_color.search(text))
        has_href = bool(pattern_link.search(text))
        return has_font_color, has_href
        
    # 使用正则表达式进行替换，这里采用了原始的正则表达式逻辑，但增加了安全转义
    pattern_color = re.compile(r'<font color=\'([0x#]+[a-fA-F0-9]{6})\'[^>]*>(.*?)</font>')
    pattern_link = re.compile(r'<a href=\'(\\S*?)\'[^>]*>(.*?)</a>')
    
    return check_patterns(html)

def ParseHtml(html):
    # 安全地处理HTML，避免XSS攻击，确保替换的内容被正确转义
    def safe_replace(match):
        color = match.group(1)
        text = match.group(2)
        # 转义text中的任何HTML特殊字符
        safe_text = re.sub(r'&|<|>', lambda m: {'&':'&amp;', '<':'&lt;', '>':'&gt;'}[m.group(0)], text)
        return f'<color={color}>{safe_text}</color>'

    def link_replace(match):
        url = match.group(1)
        text = match.group(2)
        # 转义url和text中的任何HTML特殊字符
        safe_url = re.sub(r'&|<|>', lambda m: {'&':'&amp;', '<':'&lt;', '>':'&gt;'}[m.group(0)], url)
        safe_text = re.sub(r'&|<|>', lambda m: {'&':'&amp;', '<':'&lt;', '>':'&gt;'}[m.group(0)], text)
        return f'<link="{safe_url}">{safe_text}</link>'
    
    def check_patterns(text):
        has_font_color =  bool(pattern_color.search(text))
        has_href = bool(pattern_link.search(text))
        return has_font_color, has_href

    # 使用正则表达式进行替换，这里采用了原始的正则表达式逻辑，但增加了安全转义
    pattern_color = re.compile(r'<font color=\'([0x#]+[a-fA-F0-9]{6})\'[^>]*>(.*?)</font>')
    pattern_link = re.compile(r'<a href=\'(\\S*?)\'[^>]*>(.*?)</a>')
    
    # 进行替换操作
    replaceHtml = pattern_color.sub(safe_replace, html)
    replaceHtml = pattern_link.sub(link_replace, replaceHtml)
    
    return replaceHtml

def do_excel(path):
    wb = openpyxl.load_workbook(path, False, False,True)
    sheet = wb.active
    datarow = 8
    ischang = False
    data_col_count = sheet.max_column  + 1#列数,看是否需要+1
    data_row_count = sheet.max_row
    datarow =1
    data_row_count =3
    for row_data in sheet.iter_rows(min_row=datarow , max_row=data_row_count, min_col=1, max_col=data_col_count):
        for cell in row_data:
            # 检查单元格值是否为空或非字符串类型
            if not cell.value:
                continue
            if not isinstance(cell.value, str):
                # 这里假设我们只处理字符串类型的单元格值
                continue

            # # 检查单元格值是否为自定义类型
            # if not cell.value in cusrp:
            #     continue
            # cell.value = ReplaceType(cell.value)


            a,b = HasHtml(cell.value)
            if not a and not b:
                continue
            cell.value = ParseHtml(cell.value)

            # print(cell.value)
            ischang = True

    # wb.close()
    if ischang:
        wb.save(path)
        print('change over:', path)

config.initIni()
ext = config.scriptExtDict['xlsx']
excels = config.GetFilesByExtension(config.ini['exceldir'], ext)
for excel in excels:
    # if not 'cfg_items' == excel.stem:
    #     continue
    do_excel(excel)