import copy
# from curses import raw
import re
from cv2 import DRAW_MATCHES_FLAGS_DRAW_OVER_OUTIMG
import openpyxl
import sys
import numpy as np
import config
from CellInfo import CellInfo
import NumberParse

class ExcelParse:
    ARRAY_SPLITTER = '#'
    LIST_SPLITCHAR = '\||&'
    arry32 = ['int', 'int32', 'sint32', 'sfixed32']
    arryu32 = ['uint','uint32', 'fixed32']
    arry64 =  ['int64', 'double','sint64', 'sfixed64']
    arryu64 = ['uint64', 'fixed64']
    def __init__(self, excelpath, typerow, namerow, datarow, prototype = 0, protoshow = 0):
        self.path = excelpath
        self.prototype = prototype
        self.protoshow = protoshow
        self.keytyperow = typerow
        self.keynamerow = namerow
        self.datarow = datarow
        self.isParseSuccess = False
        self.hasCommon = False
    def trimDotZeroes(self,text):
        if isinstance(text, str):
            text = text.strip()
            if text == "": return 0
            return re.sub("\.0+$", "", text)
        return text
    def get_repeate_value(self,data_type, row_value):
        arr = re.split(ExcelParse.ARRAY_SPLITTER, row_value)
        if len(arr) > 1:
            if data_type in ExcelParse.arryu64:
                v1 = int(np.int32(self.trimDotZeroes(arr[0])))
                v2 = int(np.int32(self.trimDotZeroes(arr[1])))
                return (v1<<32) | v2

        else:
            return self.get_real_value(data_type, row_value)


    def get_real_value(self,data_type, row_value):
        if not row_value :
            return None
            # print('data_type: ', data_type, 'row_value:', row_value)
        trizv = self.trimDotZeroes(row_value)
        try:
            npv = None
            if data_type == 'string':
                value = str(row_value)
                value = value.replace('\\', '\\\\')
                npv = '''{}'''.format(value)
            elif data_type in ExcelParse.arry32:
                npv = np.int32(trizv) 
            elif data_type in ExcelParse.arryu32:
                npv = np.uint32(trizv) 
            elif data_type in ExcelParse.arry64:
                npv = np.int64(trizv) 
            elif data_type in ExcelParse.arryu64:
                npv = np.uint64(trizv)
            elif data_type == 'float':
                npv = float(row_value)
            elif data_type == 'bool':
                npv = bool(row_value)
            else:
                return None
            return npv
        except ValueError as e:
            print(f"无法将 {trizv} 转换为 uint32 类型，ValueError: {e}")
            return None
    def readExcel(self, readOnly=False):
        if not config.CheckExcelFile(self.path):
            return False
        self.wb = openpyxl.load_workbook(self.path,readOnly, False,True)
        self.sheet = self.wb.active
        self.readsheet(self.sheet)
        self.wb.close()

    def readsheet(self, sheet):
        self.variableDict = [] 
        self.sheetName = sheet.title
        self.rowTableName = sheet.title
        self.groupTableName = sheet.title
        data_col_count = sheet.max_column  + 1                             #列数,看是否需要+1
        index = 1
        for col_num in range(1, data_col_count):
            name_datacell = sheet.cell(self.keynamerow, col_num)
            if not name_datacell.value:
                continue
            name_data = name_datacell.value
            if name_data == None or isinstance(name_data,str) and name_data.strip() == "": continue
            typecell = sheet.cell(self.keytyperow, col_num)
            if not typecell.value:
                continue
            type_data = typecell.value.split(':')[0]
            if type_data == None or type_data.strip() == "": continue
            variable_name = name_data	#name_data[0].upper() + name_data[1:]
            row_type_data = config.GetCustomTypeValue(type_data)
            if variable_name in self.variableDict:
                print('异常退出: ','表', self.sheetName, '存在相同的字段名: ', variable_name)
                self.isParseSuccess = False
                return 

            if not config.CheckSupportType(row_type_data):
                print('表', self.sheetName, '字段', variable_name, '的数据类型', row_type_data,'不在支持的列表中')
                self.isParseSuccess = False
                # continue
                return
            if not config.CheckDefaultType(row_type_data):
                self.hasCommon = True

            prototypecell = sheet.cell(self.prototype, col_num)
            prototype = prototypecell.value
            protoshowcell = sheet.cell(self.protoshow, col_num)
            protoshow = protoshowcell.value

            cell = CellInfo(row_type_data, type_data, variable_name, None, prototype, protoshow, index)
            # cell.protoshow = protoshow
            # cell.prototype = prototype
            cell.col = col_num

            self.variableDict.append(cell)
            index += 1

        data_row_count = sheet.max_row

        self.sheet_row_data_list = []
        for row_data in sheet.iter_rows(min_row=self.datarow , max_row=data_row_count, min_col=1, max_col=data_col_count):
            # 存储每一个字段的字段名，数值，类型
            single_row_data = []
            for variable in self.variableDict:
                rowcell = row_data[variable.col -1]
                cell = copy.deepcopy(variable)
                cell.originaldata = rowcell.value
                # if  config.CheckDefaultType(cell.type):
                #     if cell.isRepeated():
                #         if not rowcell.value == None:
                #             if '|' in rowcell.value or '&' in rowcell.value:
                #                 arrayv = re.split(ExcelParse.LIST_SPLITCHAR, rowcell.value) #rowcell.value.split(ExcelParse.ARRAY_SPLITTER)
                #                 for v in arrayv:
                #                     cell.arrayvalue.append(self.get_repeate_value(cell.type, v))
                #             else:
                #                 arrayv = re.split(ExcelParse.ARRAY_SPLITTER, rowcell.value) #rowcell.value.split(ExcelParse.ARRAY_SPLITTER)
                #                 for v in arrayv:
                #                     cell.arrayvalue.append(self.get_repeate_value(cell.type, v))
                #     else:
                #         cellvalue = self.get_real_value(cell.type, rowcell.value)
                #         cell.value = cellvalue
                # else:
                #     cell.arrayvalue = self.PareCustomType(cell.type)
                cell.arrayvalue = NumberParse.Parse(cell.typename, rowcell.value)

                cell.row = rowcell.row
                single_row_data.append(cell)
            self.sheet_row_data_list.append(single_row_data)
        self.isParseSuccess = True
