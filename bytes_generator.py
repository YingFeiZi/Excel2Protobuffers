import openpyxl
import xlrd
from datetime import date, datetime
import sys
import os
import json
import numpy as np
import re
import config

NAME_ROW = 3 
TYPE_ROW = 2
DATA_ROW = 4
ARRAY_SPLITTER = '#'

def __get_assign_code(mod_filename, mod_name, field_name, field_value, field_type, index):
	# if field_type == 'string' or field_type == '[byte]' or field_type == '[int32]' or field_type == '[int64]' or field_type == '[float]' or field_type == '[string]':
		#print("888888888",field_name)
	if field_type == 'string' or field_type in arrylist:
		value_code = """{}{}""".format(field_name, index)
	else:
		value_code = field_value
	code = _getDataModAddField( mod_filename, mod_name, field_name, value_code)
	#print(code)
	return code

def _getDataModAddField(mod_filename, mod_name, field_name, value_code):
    return f"{mod_filename}.{mod_name}Add{field_name}(builder, {value_code})"

def __getDataVectorStar(ModFileName, ModName, FieldName, ValueLen):
	return f"{ModFileName}.{ModName}Start{FieldName}Vector(builder, {ValueLen})\n"

def __getDataVectorEnd(FieldName, index):
	return f"{FieldName}{index} = builder.EndVector()\n"

def __getDataCreateString(FieldName, Index, fieldvalue, isMulti=False):
	mulstr = isMulti and "str" or ""
	return f"{FieldName}{mulstr}{Index} = builder.CreateString(\"{fieldvalue}\")\n"
def _getDataPrependByte(fieldvalue):
     return f"builder.PrependByte({fieldvalue})\n"
def _getDataPrependInt16(fieldvalue):
	return f"builder.PrependInt16({fieldvalue})\n"
def _getDataPrependUint16(fieldvalue):
	return f"builder.PrependUint16({fieldvalue})\n"
def _getDataPrependInt32(fieldvalue):
	return f"builder.PrependInt32({fieldvalue})\n"
def _getDataPrependUint32(fieldvalue):
	return f"builder.PrependUint32({fieldvalue})\n"
def _getDataPrependInt64(fieldvalue):
	return f"builder.PrependInt64({fieldvalue})\n"
def _getDataPrependUint64(fieldvalue):
	return f"builder.PrependUint64({fieldvalue})\n"
def _getDataPrependFloat(fieldvalue):
	return f"builder.PrependFloat32({fieldvalue})\n"
def _getDataPrependBool(fieldvalue):
	return f"builder.PrependBool({fieldvalue})\n"
def _getDataPrependUOffsetRelative(FieldName, index):
    return f"builder.PrependUOffsetTRelative({FieldName}Str{index})\n"


def __get_single_data_code(mod_filename, mod_name, row_data, index):
	code = \
"""
{VariableCreate}
{ModFileName}.{ModName}Start(builder)
{AssignCode}
single_data{Index} = {ModFileName}.{ModName}End(builder)
"""
	
	variable_create_code = ''
	for field in row_data:
		fvalue = field['field_value']
		if fvalue == None:
			continue
		ftype = field['field_type']
		fname = field['field_name']
  
		if ftype == 'string':
			variable_create_code += __getDataCreateString(fname, index, fvalue)
			continue
		if ftype in arrylist:
			valueArrs = fvalue.split(ARRAY_SPLITTER)
			valueArrsLen = len(valueArrs)
			subType = re.sub(r"\[|\]","", ftype)
			
			if subType == 'string':
				for curIndex in range(valueArrsLen-1,-1,-1):
					fieldvalue = valueArrs[curIndex]
					variable_create_code += __getDataCreateString(fname, curIndex,fieldvalue, True)

			variable_create_code += __getDataVectorStar(mod_filename, mod_name, fname, valueArrsLen)
			for curIndex in range(valueArrsLen-1,-1,-1):
				fieldvalue = valueArrs[curIndex]
				if subType =='string':
					variable_create_code += _getDataPrependUOffsetRelative(fname, curIndex)
				elif subType in arry8:
					variable_create_code += _getDataPrependByte(fieldvalue)
				elif subType in arry16:
					variable_create_code += _getDataPrependInt16(fieldvalue)
				elif subType in arry32:
					variable_create_code += _getDataPrependInt32(fieldvalue)
				elif subType in arry64:
					variable_create_code += _getDataPrependInt64(fieldvalue)
				elif subType in arryu8:
					variable_create_code += _getDataPrependByte(fieldvalue)
				elif subType in arryu16:
					variable_create_code += _getDataPrependUint16(fieldvalue)
				elif subType in arryu32:
					variable_create_code += _getDataPrependUint32(fieldvalue)
				elif subType in arryu64:
					variable_create_code += _getDataPrependUint64(fieldvalue)
				elif subType == 'float':
					variable_create_code += _getDataPrependFloat(fieldvalue)
				elif subType == 'bool':
					variable_create_code += _getDataPrependBool(fieldvalue)
     
			variable_create_code += __getDataVectorEnd(fname, index)

	assign_code = ''
	for field in row_data:
		if field['field_def'] == "NULL" and field['field_value'] != None:
			#print("888888",field['field_value'])
			assign_code += __get_assign_code(
											mod_filename,
											mod_name, 
											field['field_name'],
											field['field_value'],
											field['field_type'],
											index
										)
			assign_code += '\n'
		#else:
			#print("跳过不用序列化到二进制文件因为有默认值或为空",field['field_name'],field['field_value'])

	assign_code = assign_code[:-1]
	code = code.format(ModFileName = mod_filename, VariableCreate = variable_create_code, ModName = mod_name, AssignCode = assign_code, Index = index)
	return code


def __get_list_data_code(mod_name, single_mod_name, list_data):
	row_count = len(list_data)
	all_assign_code = ''
	modfilename = f"{mod_name}_generated"
	index = 0
	for row_data in list_data:
		all_assign_code += __get_single_data_code(modfilename, single_mod_name, row_data, index)
		index += 1
		all_assign_code += '\n'

	offset_code = ''
	for index in range(0, row_count):
		data_name = "single_data{}".format(index)
		offset_code += "builder.PrependUOffsetTRelative({})".format(data_name)
		offset_code += '\n'

	code = \
"""
import generated_python.{ModFileName} as {ModFileName}
import flatbuffers

builder = flatbuffers.Builder(1)

{AllAssignCode}
{ModFileName}.{ModName}StartDatalistVector(builder, {DataCount})
{OffsetCode}
data_array = builder.EndVector()

{ModFileName}.{ModName}Start(builder)
{ModFileName}.{ModName}AddDatalist(builder, data_array)
final_data = {ModFileName}.{ModName}End(builder)
builder.Finish(final_data)
buf = builder.Output()
""".format(
		# SingleModName = single_mod_name,
		ModFileName = modfilename,
		ModName = mod_name,
		AllAssignCode = all_assign_code,
		DataCount = row_count,
		OffsetCode = offset_code
	)
	
	return code


def __generate_bytes(mod_name, single_mod_name, bytes_file_root_path, excel_row_list):
	list_code = __get_list_data_code(mod_name, single_mod_name, excel_row_list)
	byte_file_path = os.path.join(bytes_file_root_path, "{}.bytes".format(mod_name))
	byte_file_path = byte_file_path.replace('\\', '/')
	code = """
{ListCode}

with open('{ByteFilePath}', 'wb') as f:
	f.write(buf)
""".format(ListCode = list_code, ByteFilePath = byte_file_path)
	# print(code)
	# if os.path.exists(f"flatc/{mod_name}.py"):
	# 	os.remove(f"{mod_name}.py")
	# file = open(f"{mod_name}.py", 'a', encoding='utf-8')
	# file.write(code)
	# file.close()
	exec(code)
	print('生成成功: ', byte_file_path)


def trimDotZeroes(text):
	if isinstance(text, str):
		text = text.strip()
		if text == "": return 0
		return re.sub("\.0+$", "", text)
	return text

arrybool = ['bool']
arry8 = ['byte']
arryu8 = ['ubyte']
arry16 = ['short', 'int16']
arryu16 = ['ushort', 'uint16']
arry32 = ['int', 'int32']
arryu32 = ['uint','uint32']
arry64 =  ['int64'] #['long', 'double']
arryu64 = ['uint64']
arrylist = ['[byte]', '[ubyte]', '[bool]', '[short]', '[ushort]', '[int]', '[uint]', '[float]', '[string]',
            '[int16]', '[uint16]', '[int32]', '[uint32]', '[int64]', '[uint64]']
def __get_real_value(data_type, raw_value):
	if not raw_value :
		return None
	# print('data_type: ', data_type, 'raw_value:', raw_value)
	if data_type == 'string':
		value = str(raw_value)
		value = value.replace('\\', '\\\\')
		return '''{}'''.format(value)
	elif data_type in arry8:
		return np.int8(trimDotZeroes(raw_value))
	elif data_type in arryu8:
		return np.uint8(trimDotZeroes(raw_value))
	elif data_type in arry16:
		return np.int16(trimDotZeroes(raw_value))
	elif data_type in arryu16:
		return np.uint16(trimDotZeroes(raw_value))
	elif data_type in arry32:
		return np.int32(trimDotZeroes(raw_value))
	elif data_type in arryu32:
		return np.uint32(trimDotZeroes(raw_value))
	elif data_type in arry64:
		return np.int64(trimDotZeroes(raw_value))
	elif data_type in arryu64:
		return np.uint64(trimDotZeroes(raw_value))
	elif data_type == 'float':
		return float(raw_value)
	elif data_type == 'bool':
		return bool(raw_value)
	elif data_type in arrylist:
		return str(raw_value)
	# elif data_type == '[int16]':
	# 	return str(raw_value)
	# elif data_type == '[uint16]':
	# 	return str(raw_value)
	# elif data_type == '[int32]':
	# 	return str(raw_value)
	# elif data_type == '[uint32]':
	# 	return str(raw_value)
	# elif data_type == '[int64]':
	# 	return str(raw_value)
	# elif data_type == '[uint64]':
	# 	return str(raw_value)
	# elif data_type == '[byte]':
	# 	return str(raw_value)
	# elif data_type == '[ubyte]':
	# 	return str(raw_value)
	# elif data_type == '[bool]':
	# 	return str(raw_value)
	# elif data_type == '[float]':
	# 	return str(raw_value)
	# elif data_type == '[string]':
	# 	return str(raw_value)
	else:
		return None

def __read_excel_sheet(sheet):
	variable_dict = {}
	variable_index = {}
	variable_defaultValue_dict = {}
	# sheet_name = sheet.name
	sheet_name = sheet.title
	mod_name = sheet_name
	single_mod_name = mod_name + 'RowData'
	# row_table_name = sheet_name + 'RowData'
	# group_table_name = sheet_name;
	# data_col_count = sheet.ncols#列数
	data_col_count = sheet.max_column#列数
	for col_num in range(1, data_col_count):
		namecell  = sheet._get_cell(NAME_ROW, col_num)
		if not namecell.value:
			continue
		name_data = str(namecell.value)
		if name_data == None or name_data.strip() == "": continue

		if name_data =='getway':
			pass

		typecell  = sheet._get_cell(TYPE_ROW, col_num)
		if not typecell.value:
			continue
		type_data = str(typecell.value).split(':')
		if name_data == None or type_data[0].strip() == "": continue

		variable_name = name_data[0].upper() + name_data[1:]
		row_type_data = config.GetCustomTypeValue(type_data[0])
		if variable_name in variable_dict:
			print('存在相同的字段名: ', variable_name)
			print('异常退出')
			sys.exit()

		if not config.CheckSupportType(row_type_data):
			#print('字段', variable_name, '的数据类型', row_type_data,'不在支持的列表中')
			#print('异常退出')
			#sys.exit()
			continue
		variable_dict[variable_name] = row_type_data
		variable_index[variable_name] = col_num
		
		typeLen = len(type_data)
		if typeLen == 2:
			variable_defaultValue_dict[variable_name] = type_data[1]
			print(variable_name,"此处有默认值")
		else:
			variable_defaultValue_dict[variable_name] = "NULL"

	data_row_count = sheet.max_row

	sheet_row_data_list = []
	#for x in range(1, data_row_count):
	# for x in range(data_row_count-1, DATA_ROW-1,-1):
	for row_data in sheet.iter_rows(min_row=DATA_ROW, max_row=data_row_count, min_col=1, max_col=data_col_count):
		# row_data = sheet.row(x)
		# 存储每一个字段的字段名，数值，类型
		single_row_data = []
		
		for variable_name in variable_dict:
			variable_type = variable_dict[variable_name]
			index = variable_index[variable_name]
			#print(variable_name, variable_type, row_data[index].value)
			variable_value = __get_real_value(variable_type, row_data[index -1].value)
			variable_def_calue = variable_defaultValue_dict[variable_name]
			if variable_def_calue == "NULL":
				variable_def_calue = variable_defaultValue_dict[variable_name]
			else:
				variable_def_calue = __get_real_value(variable_type, variable_def_calue)
				if variable_def_calue != variable_value:
					variable_def_calue = "NULL"
			# print(variable_name, variable_type, variable_value)
			index += 1
			data_dict = {
				'field_name': variable_name,
				'field_value': variable_value,
				'field_type': variable_type,
				'field_def': variable_def_calue
			}
			single_row_data.append(data_dict)
		sheet_row_data_list.append(single_row_data)
	__generate_bytes(mod_name, single_mod_name, config.GetRootBytes(), sheet_row_data_list)


def __generate_excel_data(excel_path):
	# wb = xlrd.open_workbook(excel_path)
	# sheet_count = 1#len(wb.sheet_names())
	# sheet1 = wb.sheet_by_index(0)
	# for x in range(0, sheet_count):
	# 	sheet = wb.sheet_by_index(x)
	# 	__read_excel_sheet(sheet)
	# print("处理文件:",excel_path)
	wb = openpyxl.load_workbook(excel_path,True, False,True)
	sheet = wb.active
	__read_excel_sheet(sheet)


def __generate_all_excel_byte_data():
	excel_root = config.GetRootExcel()
	for root, dirs, files in os.walk(excel_root):
		for file in files:
			excel_file_path = os.path.join(root, file)
			if config.CheckExcelFile(excel_file_path) and not file.startswith('~'):
				__generate_excel_data(excel_file_path)

"""
def __read_csv_sheet(sheet,fileName):
	variable_dict = {}
	variable_defaultValue_dict = {}
	sheet_name = fileName
	mod_name = sheet_name
	single_mod_name = mod_name + 'RowData'
	#print("文件名",fileName)
	result = list(sheet)
	row1 = result[0]#第一行
	row2 = result[1]#第二行
	row3 = result[2]#第三行
	data_col_count = len(row1)#列数
	for i in range(data_col_count):
		name_data = row1[i]
		type_data = row2[i].split(':')
		variable_name = name_data[0].upper() + name_data[1:]
		row_type_data = type_data[0]
		if variable_name in variable_dict:
			print('存在相同的字段名: ', variable_name)
			print('异常退出')
			sys.exit()

		if not config.CheckSupportType(row_type_data):
			print('字段', variable_name, '的数据类型', row_type_data,'不在支持的列表中')
			print('异常退出')
			sys.exit()
		variable_dict[variable_name] = row_type_data
		
		typeLen = len(type_data)
		if typeLen == 2:
			variable_defaultValue_dict[variable_name] = type_data[1]
			print(variable_name,"此处有默认值")
		else:
			variable_defaultValue_dict[variable_name] = "NULL"
	data_row_count = len(result)#行数
	#print("列数和行数",data_col_count,data_row_count)
	sheet_row_data_list = []
	#for x in range(1, data_row_count):
	for x in range(data_row_count-1, 2,-1):
		row_data = result[x]
		# 存储每一个字段的字段名，数值，类型
		single_row_data = []
		
		index = 0
		for variable_name in variable_dict:
			variable_type = variable_dict[variable_name]
			#print("字段",variable_name, variable_type, row_data[index])
			variable_value = __get_real_value(variable_type, row_data[index])
			variable_def_calue = variable_defaultValue_dict[variable_name]
			if variable_def_calue == "NULL":
				variable_def_calue = variable_defaultValue_dict[variable_name]
			else:
				variable_def_calue = __get_real_value(variable_type, variable_def_calue)
				if variable_def_calue != variable_value:
					variable_def_calue = "NULL"
			# print(variable_name, variable_type, variable_value)
			index += 1
			data_dict = {
				'field_name': variable_name,
				'field_value': variable_value,
				'field_type': variable_type,
				'field_def': variable_def_calue
			}
			single_row_data.append(data_dict)
		sheet_row_data_list.append(single_row_data)
	__generate_bytes(mod_name, single_mod_name, config.GetRootBytes(), sheet_row_data_list)

def __generate_csv_data(excel_path,fileName):
	with open(excel_path, 'r',encoding='utf-8') as f:
		reader = csv.reader(f)
		# header_row = csv.reader(f)
		__read_csv_sheet(reader,fileName)


def __generate_all_csv_byte_data():
	for root, dirs, files in os.walk(config.GetRootExcel()):
		for file in files:
			excel_file_path = os.path.join(root, file)
			if config.CheckCsvFile(excel_file_path) and not file.startswith('~'):
				filename = file.split('.')
				__generate_csv_data(excel_file_path,filename[0])
"""

def run():
	print('---------------- 将excel生成flatbuffers二进制数据 ----------------')
	__generate_all_excel_byte_data()
	# __generate_all_csv_byte_data()