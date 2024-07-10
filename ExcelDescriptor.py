import copy
# from curses import raw
import re
from cv2 import DRAW_MATCHES_FLAGS_DRAW_OVER_OUTIMG
import openpyxl
from google.protobuf.descriptor import FieldDescriptor
from regex import E
import Encrypted
from io import BytesIO
from google.protobuf.internal.encoder import _EncodeVarint
from google.protobuf.message import Message
import sys
from pathlib import Path
import numpy as np
import config
from CellInfo import CellInfo
import table_common_pb2
import NumberParse


class ExcelDescriptor:

    def __init__(self, excelpath, typerow, namerow, datarow, prototype = 0, protoshow = 0):
        self.path = excelpath
        self.prototype = prototype
        self.protoshow = protoshow
        self.keytyperow = typerow
        self.keynamerow = namerow
        self.datarow = datarow
        self.isParseSuccess = False
        self.hasCommon = False
   
    def PareProtoDesc(self, pb2, entry_name):
        module = __import__(pb2)
        self.proto_desc = []
        # alias FieldDescriptor
        # FieldDescriptor = google.protobuf.descriptor.FieldDescriptor
        self.entry_name = getattr(module, entry_name)
        DESCRIPTOR = self.entry_name.DESCRIPTOR
        for desc in DESCRIPTOR.fields:
            if desc.number > 99:
                continue
            field_name = desc.name
            if field_name[0] == '_':
                continue

            field_number = desc.number
            field_label = desc.label
            field_type = None
            if desc.type == desc.TYPE_INT32 or \
                    desc.type == desc.TYPE_SINT32 or \
                    desc.type == desc.TYPE_FIXED32 or \
                    desc.type == desc.TYPE_UINT32 or \
                    desc.type == desc.TYPE_ENUM :
                field_type = np.int32
            if desc.type == desc.TYPE_INT64 or \
                    desc.type == desc.TYPE_SINT64 or \
                    desc.type == desc.TYPE_FIXED64 or \
                    desc.type == desc.TYPE_UINT64:
                field_type = np.int64
            elif desc.type == desc.TYPE_BOOL:
                field_type = bool
            elif desc.type == desc.TYPE_FLOAT:
                field_type = float
            elif desc.type == desc.TYPE_BYTES or \
                    desc.type == desc.TYPE_STRING:
                field_type = str
            elif desc.type == desc.TYPE_MESSAGE:
                msg_desc = desc.message_type
                field_type = msg_desc.name
                if msg_desc.name == "Int64List":
                    field_type = table_common_pb2.Int64List
                elif msg_desc.name == "Int32List":
                    field_type = table_common_pb2.Int32List
                elif msg_desc.name == "Int32ListList":
                    field_type = table_common_pb2.Int32ListList
                else:
                    print("bad field type: " + msg_desc.name)

            self.proto_desc += [(field_name, field_number, field_type, field_label)]
        # print("proto_desc = " ,str(self.proto_desc))


    def readExcel(self, readOnly=True):
        if not config.CheckExcelFile(self.path):
            return False
        self.wb = openpyxl.load_workbook(self.path,readOnly, False,True)
        self.sheet = self.wb.active
        self.readsheet(self.sheet)
        self.wb.close()

    def readsheet(self, sheet):
        self.variableDict = [] 
        self.sheetName = Path(self.path).stem
        self.pbName = f"{self.sheetName}_pb2"
        self.rowTableName = self.sheetName
        self.groupTableName = self.sheetName
        data_col_count = sheet.max_column  + 1                             #列数,看是否需要+1

        self.PareProtoDesc(self.pbName, self.sheetName.upper())
        self.row_array = []
        data_row_count = sheet.max_row

        ENUM_FIELD_NAME = 0
        ENUM_FIELD_NUMBER = 1
        ENUM_FIELD_TYPE = 2
        ENUM_FIELD_LABEL = 3


        rowCount = self.datarow
        for row_data in sheet.iter_rows(min_row=self.datarow , max_row=data_row_count, min_col=1, max_col=data_col_count):
            # print("row_values = " + str(row_data))
            # 存储每一个字段的字段名，数值，类型
            row = self.entry_name()
            isadd = False
            for field_desc in self.proto_desc:
                # print(field_desc)
                celCount = field_desc[ENUM_FIELD_NUMBER]-1
                row_value = row_data[celCount].value
                # print("SSS" + str(field_desc[ENUM_FIELD_NUMBER]-1))
                # print(str(row))
                # print("---------------1------------------------")
                if row_value == None or row_value == '':
                    continue
                isadd = True

                try:
                    if type(row_value) == float:
                        row_value = str(row_value)

                    if field_desc[ENUM_FIELD_TYPE] == str:
                        if field_desc[ENUM_FIELD_LABEL] == FieldDescriptor.LABEL_REQUIRED:
                            if type(row_value) != type(u"1") and type(row_value) != type("1"):
                                self.LogTableInfo(row_data.row, field_desc[ENUM_FIELD_NUMBER])
                            setattr(row, field_desc[ENUM_FIELD_NAME], row_value)
                        elif field_desc[ENUM_FIELD_LABEL] == FieldDescriptor.LABEL_REPEATED:
                            vlist =  NumberParse.ParseStringList(row_value)
                            for section in vlist:
                                section = section.strip()
                                if section != "":
                                    getattr(row, field_desc[ENUM_FIELD_NAME]).append(section)
                        else:
                            setattr(row, field_desc[ENUM_FIELD_NAME], str(row_value))

                    # int32
                    elif field_desc[ENUM_FIELD_TYPE] == np.int32:
                        if field_desc[ENUM_FIELD_LABEL] == FieldDescriptor.LABEL_REQUIRED:
                            if type(row_value) != type("1"):
                                self.LogTableInfo(row_data.row, field_desc[ENUM_FIELD_NUMBER])
                            if self.CheckNoneOrNull(row_value):
                                row_value = int(0)
                            else:
                                row_value = int(row_value)
                            setattr(row, field_desc[ENUM_FIELD_NAME], row_value)
                        elif field_desc[ENUM_FIELD_LABEL] == FieldDescriptor.LABEL_REPEATED:
                            vlist = NumberParse.ParseStringToIntList(row_value)
                            for section in  vlist:
                                if not section == None:
                                    getattr(row, field_desc[ENUM_FIELD_NAME]).append(section)

                        else:
                            if self.CheckNoneOrNull(row_value):
                                row_value = int(0)
                            else:
                                row_value = int(row_value)

                            setattr(row, field_desc[ENUM_FIELD_NAME], row_value)
                    
                    elif field_desc[ENUM_FIELD_TYPE] == np.int64:
                        # if field_desc[ENUM_FIELD_LABEL] == FieldDescriptor.LABEL_REQUIRED:
                        #     if type(row_value) != type("1"):
                        #         self.LogTableInfo(row_data.row, field_desc[ENUM_FIELD_NUMBER])
                        #     if self.CheckNoneOrNull(row_value):
                        #         row_value = int(0)
                        #     else:
                        #         row_value = int(row_value)
                        #     setattr(row, field_desc[ENUM_FIELD_NAME], row_value)
                        # el
                        if field_desc[ENUM_FIELD_LABEL] == FieldDescriptor.LABEL_REPEATED:
                            valueList = []
                            if not self.CheckNoneOrNull(row_value):
                                if config.LIST_SPLITCHAR1 in row_value :
                                    valueList = NumberParse.ParseStringToComboList(row_value, config.LIST_SPLITCHAR1, config.ARRAY_SPLITTER)
                                elif config.LIST_SPLITCHAR2 in row_value:
                                    valueList = NumberParse.ParseStringToComboList(row_value, config.LIST_SPLITCHAR2, config.ARRAY_SPLITTER)
                            for section in  valueList:
                                getattr(row, field_desc[ENUM_FIELD_NAME]).append(section)

                        else:
                            if self.CheckNoneOrNull(row_value):
                                row_value = int(0)
                            elif NumberParse.CheckSplitChar(row_value):
                                row_value = NumberParse.ParseStringToCombo(row_value)
                            else:
                                row_value = int(row_value)
                            setattr(row, field_desc[ENUM_FIELD_NAME], row_value)

                    elif field_desc[ENUM_FIELD_TYPE] == bool:
                        if field_desc[ENUM_FIELD_LABEL] == FieldDescriptor.LABEL_REQUIRED:
                            if type(row_value) != type(True):
                                self.LogTableInfo(row_data.row, field_desc[ENUM_FIELD_NUMBER])
                            setattr(row, field_desc[ENUM_FIELD_NAME], bool(row_value))
                        elif field_desc[ENUM_FIELD_LABEL] == FieldDescriptor.LABEL_REPEATED:
                            vlist = NumberParse.ParseStringToIntList(row_value)
                            for section in vlist:
                                if section != "":
                                    getattr(row, field_desc[ENUM_FIELD_NAME]).append(bool(section))
                        else:
                            if self.CheckNoneOrNull(row_value):
                                row_value = bool(0)
                            else:
                                row_value = bool(row_value)
                            setattr(row, field_desc[ENUM_FIELD_NAME], row_value)
                    #float
                    elif field_desc[ENUM_FIELD_TYPE] == float:
                        if field_desc[ENUM_FIELD_LABEL] == FieldDescriptor.LABEL_REQUIRED:
                            if self.CheckNoneOrNull(row_value):
                                row_value = float(0)
                            else:
                                row_value = float(row_value)
                            setattr(row,field_desc[ENUM_FIELD_NAME], row_value)
                        elif field_desc[ENUM_FIELD_LABEL] == FieldDescriptor.LABEL_REPEATED:
                            vlist =  NumberParse.ParseStringList(row_value)
                            for section in vlist:
                                if section != "":
                                    getattr(row,field_desc[ENUM_FIELD_NAME].append(float(section)))
                        else:
                            if self.CheckNoneOrNull(row_value):
                                row_value = float(0)
                            else:
                                row_value = float(row_value)
                            setattr(row,field_desc[ENUM_FIELD_NAME], row_value)

                    # Int64List
                    elif field_desc[ENUM_FIELD_TYPE] == table_common_pb2.Int64List:
                        if not self.CheckNoneOrNull(row_value):
                            for data in NumberParse.ParseInt64List(row_value):
                                item = self.ParseInt64List(data)
                                if not item:
                                    print("pass item drop failed: " + row_value)
                                getattr(row, field_desc[ENUM_FIELD_NAME]).append(item)

                    # Int32List
                    elif field_desc[ENUM_FIELD_TYPE] == table_common_pb2.Int32List:
                        if not self.CheckNoneOrNull(row_value):
                            for data in NumberParse.ParseInt32List(row_value):
                                item = self.ParseInt32List(data)
                                if not item:
                                    print("pass item drop failed: " + row_value)
                                getattr(row, field_desc[ENUM_FIELD_NAME]).append(item)
                    
                    # Int32ListList
                    elif field_desc[ENUM_FIELD_TYPE] == table_common_pb2.Int32ListList:
                        if not self.CheckNoneOrNull(row_value):
                            for data in NumberParse.ParseInt32ListList(row_value):
                                item = self.ParseInt32ListList(data)
                                if not item:
                                    print("pass item drop failed: " + row_value)
                                getattr(row, field_desc[ENUM_FIELD_NAME]).append(item)
                    

                    else:
                        print("invalid field type: " + str(field_desc[ENUM_FIELD_TYPE]))
                except Exception as e:
                    print(f"<br><font color='red'>{self.sheetName}, row:{rowCount}  cel:{celCount} error: {e}</font>")
            if isadd:
                self.row_array.append(row)
            rowCount += 1

        self.isParseSuccess = True

    def CheckNoneOrNull(self,value):
        return value == None or value == ''  or value == ""

    def Write(self, path):
        with open(path, 'wb') as f:
            b = BytesIO()
            element_count = int(len(self.row_array))
            _EncodeVarint(b.write, element_count)
            for message in self.row_array:
                if isinstance(message, Message):  # 确保是protobuf消息对象
                    # 计算消息序列化后的字节大小
                    size = message.ByteSize()
                    _EncodeVarint(b.write, size)
                    b.write(message.SerializeToString())
            bytes_to_write = b.getvalue()
            dataencrypt = Encrypted.xor_cipher_default(bytes_to_write)
            f.write(dataencrypt)
            b.close()

    def LogTableInfo(self, row,col):
        print(f"<br>数据类型错误，在表格 {str(row)} row {str(col) }col")

    def ParseInt64List(self,row_value):
        VList= table_common_pb2.Int64List()
        for data in row_value:
            VList.list.append(data)
        return VList
    
    def ParseInt32List(self,row_value):
        VList= table_common_pb2.Int32List()
        for data in row_value:
            VList.list.append(data)
        return VList
    
    def ParseInt32ListList(self,row_value):
        VList= table_common_pb2.Int32ListList()
        for data in row_value:
            VList.list.append(self.ParseInt32List(data))
        return VList